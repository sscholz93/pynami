# -*- coding: utf-8 -*-
"""
Some utility functions for user convenience
"""
# Standard library imports
import io
import os
import csv
import webbrowser
import datetime
from tkinter import Tk
from tkinter.filedialog import asksaveasfilename

# Third party imports
from tabulate import tabulate
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


def send_emails(mitglieder, to='', method='bcc', email1=True, email2=True,
                open_browser=True):
    """
    Send emails to several members.

    Args:
        mitglieder (list): The List contents can be either
            :class:`~pynami.schemas.SearchMitglied` or
            :class:`~pynami.schemas.Mitglied`
        to (:obj:`str`, optional): Primary recipient
        method (:obj:`str`, optional): If you want to send your mails as bcc
            or something else. Currently only bcc is supported.
        email1 (:obj:`bool`, optional): If emails should be send to the
            primary address of the members.
        email2 (:obj:`bool`, optional): If emails should be send to the email
            account of the member's parent.
        open_browser (:obj:`bool`, optional): If :data:`True` the link is
            opened directly by the system. On a computer this may open your
            default mail program.

    Returns:
        str: The mailto link
    """
    recipients = []
    if email1:
        recipients += [mgl.email for mgl in mitglieder if mgl.email]
    if email2:
        recipients += [mgl.emailVertretungsberechtigter for mgl in mitglieder
                       if mgl.emailVertretungsberechtigter]
    url = f"mailto:{to}?{method}={','.join(set(recipients))}"
    if open_browser:
        webbrowser.open(url, new=1)
    return url


def tabulate2x(objs, elements=None):
    """
    Tabulate a list of objects by tabulating each object first

    Args:
        obj (list): The list of objects to tabulate. If they are not from the
            same class this may not work.
        elements (:obj:`list` of :obj:`str`, optional): List of keys which
            should be displayed

    Returns:
        str: Nicely formatted tabulated output
    """
    return tabulate([x.tabulate(elements=elements) for x in objs],
                    headers='keys')


def make_csv(data, attrs=None, includeheader=True, delimiter=','):
    """
    Makes a |CSV| formatted string from a data set

    Args:
        data (list): Data objects. They should all belong to the same class.
        attrs (:obj:`list` of `str`, optional): Attribute names for the |CSV|
            table. If left empty (:data:`None`) the value of the first
            :attr:`~.schemas.base.BaseModel._tabkeys` attribute in the list is
            taken.
        includeheader (:obj:`bool`, optional): Whether to include headers in
            the output. Defaults to :data:`True`.

    Returns:
        str: |CSV| formatted data
    """
    if not data:
        return ''
    if not attrs:
        attrs = data[0]._tabkeys
    data = [x.tabulate(attrs) for x in data]
    output = io.StringIO()
    w = csv.DictWriter(output, attrs, quoting=csv.QUOTE_NONNUMERIC,
                       delimiter=delimiter)
    if includeheader:
        w.writeheader()
    w.writerows(data)
    return output.getvalue()


def export_xlsx(data, attrs=None, includeheader=True, tableName='Tabelle1',
                sheetName='Data', write_to_file=False, filepath=''):
    """
    Create a Microsoft Excel Wokbook from a given dataset. The data can
    optionally be saved to a file.

    Args:
        data (list): Data objects. They should all belong to the same class.
        attrs (:obj:`list` of `str`, optional): Attribute names for the |CSV|
            table. If left empty (:data:`None`) the value of the first
            :attr:`~.schemas.base.BaseModel._tabkeys` attribute in the
            list is taken.
        includeheader (:obj:`bool`, optional): Whether to include headers in
            the output. Defaults to :data:`True`.
        tableName (:obj:`str`, optional): Name of the table. Defaults to
            `'Tabelle1'`.
        sheetName (:obj:`str`, optional): Name of the worksheet. Defaults to
            `'Data'`.
        write_to_file (:obj:`bool`, optional): If the workbook should be saved
            to a file. Defaults to :data:`False`.
        filepath (:obj:`str`, optional): Full path to the Excel file where the
            data should be saved. If left empty the default savefile dialog will
            be invoked via the module :mod:`tkinter`.

    Returns:
        :class:`~openpyxl.workbook.workbook.Workbook`: The created workbook.
    """
    wb = Workbook()
    if not data:
        return wb
    # Get headings and format data
    if not attrs:
        attrs = data[0]._tabkeys
    data = [[getattr(x, a) for a in attrs] for x in data]

    # Write data to worksheet
    ws = wb.active
    ws.title = sheetName
    if includeheader:
        ws.append(attrs)
    for row in data:
        ws.append(row)
    for row in ws:
        for cell in row:
            if isinstance(cell.value, (datetime.datetime, datetime.date)):
                cell.value = cell.value.strftime('%d.%m.%Y')
                cell.number_format = 'dd.mm.yyyy'

    # Create the table
    tab = Table(displayName=tableName,
                ref='A1:' + get_column_letter(ws.max_column) + str(ws.max_row))
    style = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True,
                           showColumnStripes=False)
    tab.tableStyleInfo = style

    # Add the table to the worksheet
    ws.add_table(tab)

    # Optional saving as file
    if write_to_file:
        if not filepath:
            Tk().withdraw()
            filepath = asksaveasfilename(filetypes=[('Excel files', '*.xlsx')],
                                         initialdir = os.getcwd(),
                                         defaultextension=".xlsx")
        if filepath:
            wb.save(filepath)

    # Return the workbook
    return wb
